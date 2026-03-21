"""
K-Trader Master v7.0 - 데이터베이스 모듈
[개선 #1] IPC 소켓 교체 완료 (폴링 테이블 제거)
[개선 #2] 순수 매매 기록(history) 및 통계 보관용으로 경량화
[개선 #3] 웹 모니터 및 종료 리포트 지원 메서드 추가
[개선 #4] 엔진 상태 파일(JSON) 기반 공유 메커니즘 추가
"""
import os
import sqlite3
import json
import logging
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger("ktrader")


class Database:
    """SQLite WAL 모드 DB 관리자."""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = None
        self._connect()
        self._init_tables()

        # 엔진 상태 공유 파일 (웹 모니터용)
        self._state_file = os.path.join(os.path.dirname(db_path), "engine_state.json")

    def _connect(self):
        try:
            # [v10.5.1/M6] check_same_thread=False: 현재 엔진은 PyQt5 메인 스레드에서만
            # DB 접근하므로 문제 없으나, IPC 스레드 등에서 DB 접근 시 데이터 손상 위험.
            # 멀티스레드 접근이 필요하면 스레드별 connection을 생성하거나 큐 기반으로 전환할 것.
            self.conn = sqlite3.connect(self.db_path, timeout=10, isolation_level=None, check_same_thread=False)
            self.conn.execute("PRAGMA journal_mode=WAL")
            self.conn.execute("PRAGMA busy_timeout=5000")
            logger.info("✅ [DB] SQLite 데이터베이스 연동 성공 (로깅 모드)")
        except sqlite3.Error as e:
            logger.critical(f"❌ [DB] 데이터베이스 연결 실패: {e}")
            raise

    def _safe_execute(self, query, params=None, is_write=True):
        """
        [v10.6 안정성] 자동 재연결 포함 안전한 쿼리 실행.
        장시간 가동 시 디스크 I/O 에러나 WAL 체크포인트 실패로
        연결이 깨질 수 있으므로, 1회 재연결 후 재시도합니다.
        """
        for attempt in range(2):
            try:
                cursor = self.conn.cursor()
                if params:
                    cursor.execute(query, params)
                else:
                    cursor.execute(query)
                return cursor
            except (sqlite3.OperationalError, sqlite3.DatabaseError) as e:
                if attempt == 0:
                    logger.warning(f"⚠️ [DB] 쿼리 실행 실패 → 재연결 시도: {e}")
                    try:
                        if self.conn:
                            self.conn.close()
                    except Exception:
                        pass
                    try:
                        self._connect()
                    except Exception as reconnect_err:
                        logger.error(f"❌ [DB] 재연결 실패: {reconnect_err}")
                        raise e  # 원래 에러를 raise
                else:
                    raise

    def _init_tables(self):
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS trade_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    time TEXT NOT NULL,
                    trade_type TEXT NOT NULL,
                    condition_name TEXT,
                    stock_name TEXT,
                    stock_code TEXT,
                    exec_price INTEGER DEFAULT 0,
                    exec_qty INTEGER DEFAULT 0,
                    realized_profit INTEGER DEFAULT 0,
                    commission INTEGER DEFAULT 0,
                    tax INTEGER DEFAULT 0,
                    order_type TEXT DEFAULT '시장가',
                    is_mock INTEGER DEFAULT 0,
                    sell_reason TEXT DEFAULT ''
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS daily_summary (
                    date TEXT PRIMARY KEY,
                    total_trades INTEGER DEFAULT 0,
                    wins INTEGER DEFAULT 0,
                    losses INTEGER DEFAULT 0,
                    realized_profit INTEGER DEFAULT 0,
                    max_drawdown INTEGER DEFAULT 0,
                    avg_hold_seconds REAL DEFAULT 0
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS condition_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    time TEXT NOT NULL,
                    stock_code TEXT,
                    stock_name TEXT,
                    cond_name TEXT,
                    result TEXT,
                    reason TEXT
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS blacklist_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    time TEXT NOT NULL,
                    action TEXT NOT NULL,
                    stock_code TEXT,
                    stock_name TEXT,
                    reason TEXT
                )
            ''')
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_cond_log_date ON condition_log(date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_bl_log_date ON blacklist_log(date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_trade_date ON trade_history(date)")

            # 기존 DB 마이그레이션 — 컬럼 누락 시 추가
            existing = [row[1] for row in cursor.execute("PRAGMA table_info(trade_history)").fetchall()]
            for col, definition in [("is_mock", "INTEGER DEFAULT 0"), ("sell_reason", "TEXT DEFAULT ''")]:
                if col not in existing:
                    cursor.execute(f"ALTER TABLE trade_history ADD COLUMN {col} {definition}")
                    logger.info(f"✅ [DB] trade_history.{col} 컬럼 추가 (마이그레이션)")
            logger.info("✅ [DB] 로깅 테이블 초기화 완료")
        except sqlite3.Error as e:
            logger.critical(f"❌ [DB] 테이블 생성 실패: {e}")
            raise

    # ── 조건식 편입 로그 ──────────────────────────────
    def log_condition_signal(self, code: str, name: str, cond_name: str, result: str, reason: str = ""):
        """조건식 편입 신호를 DB에 영구 저장."""
        try:
            now = datetime.datetime.now()
            self._safe_execute(
                """INSERT INTO condition_log
                   (date, time, stock_code, stock_name, cond_name, result, reason)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
                 code, name, cond_name, result, reason)
            )
        except (sqlite3.Error, Exception) as e:
            logger.error(f"❌ [DB] 조건식 로그 저장 실패: {e}")

    def get_condition_log(self, date: str = None, limit: int = 500) -> list:
        """조건식 편입 로그 조회."""
        if date is None:
            date = datetime.datetime.now().strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                """SELECT time, stock_name, stock_code, cond_name, result, reason
                   FROM condition_log WHERE date = ? ORDER BY rowid DESC LIMIT ?""",
                (date, limit)
            )
            return cursor.fetchall()
        except sqlite3.Error:
            return []

    # ── 블랙리스트 기록 ──────────────────────────────
    def log_blacklist(self, action: str, code: str, name: str, reason: str = ""):
        """블랙리스트 추가/제거 이력 저장."""
        try:
            now = datetime.datetime.now()
            self._safe_execute(
                """INSERT INTO blacklist_log
                   (date, time, action, stock_code, stock_name, reason)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
                 action, code, name, reason)
            )
        except (sqlite3.Error, Exception) as e:
            logger.error(f"❌ [DB] 블랙리스트 로그 저장 실패: {e}")

    def get_blacklist_log(self, days: int = 30) -> list:
        """블랙리스트 이력 조회."""
        cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                """SELECT date, time, action, stock_code, stock_name, reason
                   FROM blacklist_log WHERE date >= ? ORDER BY rowid DESC""",
                (cutoff,)
            )
            return cursor.fetchall()
        except sqlite3.Error:
            return []

    # ── 매매 기록 ──────────────────────────────────
    def log_trade(self, trade_type, cond_name, name, code, price, qty, realized=0, commission=0, tax=0, order_type="시장가", is_mock=False, sell_reason=""):
        try:
            now = datetime.datetime.now()
            today = now.strftime("%Y-%m-%d")
            t_str = now.strftime("%H:%M:%S")

            self._safe_execute(
                """INSERT INTO trade_history
                   (date, time, trade_type, condition_name, stock_name, stock_code,
                   exec_price, exec_qty, realized_profit, commission, tax, order_type, is_mock, sell_reason)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (today, t_str, trade_type, cond_name, name, code, price, qty, realized, commission, tax, order_type, int(is_mock), sell_reason)
            )
            logger.info(f"📝 [DB] 거래 기록: {trade_type} {name}({code}) {qty}주 @{price:,} 손익={realized:+,} 수수료={commission:,} 세금={tax:,}")
        except (sqlite3.Error, Exception) as e:
            logger.error(f"❌ [DB] 거래 기록 실패: {e}")

    # ── 매매 통계 ──────────────────────────────────
    def get_today_trades(self):
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute(
                "SELECT time, trade_type, condition_name, stock_name, exec_price, exec_qty, realized_profit "
                "FROM trade_history WHERE date = ? ORDER BY id", (today,)
            )
            return cursor.fetchall()
        except sqlite3.Error:
            return []

    def get_today_trade_summary(self) -> dict:
        """
        [개선 #3] 오늘의 매매 통계 요약 (종료 리포트용).
        반환: {'buy_count', 'buy_amount', 'sell_count', 'sell_amount', 'wins', 'losses', 'realized_profit'}
        """
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()

            # 매수 통계
            cursor.execute(
                "SELECT COUNT(*), COALESCE(SUM(exec_price * exec_qty), 0) "
                "FROM trade_history WHERE date = ? AND trade_type = '매수'", (today,)
            )
            buy_row = cursor.fetchone()
            buy_count = buy_row[0] if buy_row else 0
            buy_amount = buy_row[1] if buy_row else 0

            # 매도 통계
            cursor.execute(
                """SELECT COUNT(*),
                    COALESCE(SUM(exec_price * exec_qty), 0),
                    SUM(CASE WHEN realized_profit > 0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN realized_profit <= 0 THEN 1 ELSE 0 END),
                    COALESCE(SUM(realized_profit), 0)
                FROM trade_history WHERE date = ? AND trade_type = '매도'""", (today,)
            )
            sell_row = cursor.fetchone()
            sell_count = sell_row[0] if sell_row else 0
            sell_amount = sell_row[1] if sell_row else 0
            wins = sell_row[2] if sell_row and sell_row[2] else 0
            losses = sell_row[3] if sell_row and sell_row[3] else 0
            realized = sell_row[4] if sell_row else 0

            return {
                "buy_count": buy_count,
                "buy_amount": buy_amount,
                "sell_count": sell_count,
                "sell_amount": sell_amount,
                "wins": wins,
                "losses": losses,
                "realized_profit": realized,
            }
        except sqlite3.Error as e:
            logger.error(f"❌ [DB] 오늘 통계 조회 실패: {e}")
            return {
                "buy_count": 0, "buy_amount": 0, "sell_count": 0,
                "sell_amount": 0, "wins": 0, "losses": 0, "realized_profit": 0,
            }

    def get_statistics(self, days=30):
        cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT
                    COUNT(*),
                    SUM(CASE WHEN realized_profit > 0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN realized_profit < 0 THEN 1 ELSE 0 END),
                    SUM(CASE WHEN realized_profit = 0 THEN 1 ELSE 0 END),
                    COALESCE(SUM(realized_profit), 0),
                    COALESCE(AVG(realized_profit), 0),
                    COALESCE(MAX(realized_profit), 0),
                    COALESCE(MIN(realized_profit), 0),
                    COALESCE(AVG(CASE WHEN realized_profit > 0 THEN realized_profit END), 0),
                    COALESCE(AVG(CASE WHEN realized_profit < 0 THEN realized_profit END), 0)
                FROM trade_history WHERE date >= ? AND trade_type = '매도'
            """, (cutoff,))
            row = cursor.fetchone()
            if not row or row[0] == 0:
                return None
            total, wins, losses, be, total_profit, avg_profit, best, worst, avg_win, avg_loss = row
            return {
                "total_sells": total, "wins": wins, "losses": losses, "breakeven": be,
                "win_rate": round((wins / total * 100) if total else 0, 1),
                "total_profit": total_profit, "avg_profit": int(avg_profit),
                "best_trade": best, "worst_trade": worst,
                "avg_win": int(avg_win), "avg_loss": int(avg_loss),
                "profit_factor": round(abs(avg_win / avg_loss), 2) if avg_loss else float('inf')
            }
        except sqlite3.Error:
            return None

    def get_condition_performance(self, days=30):
        cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT condition_name, COUNT(*),
                    SUM(CASE WHEN realized_profit > 0 THEN 1 ELSE 0 END), SUM(realized_profit)
                FROM trade_history WHERE date >= ? AND trade_type = '매도' GROUP BY condition_name
            """, (cutoff,))
            return cursor.fetchall()
        except sqlite3.Error:
            return []

    def get_daily_pnl(self, days=30):
        """[개선 #3] 일별 손익 조회 (웹 모니터 차트용)."""
        cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
        try:
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT date, COALESCE(SUM(realized_profit), 0), COUNT(*)
                FROM trade_history
                WHERE date >= ? AND trade_type = '매도'
                GROUP BY date ORDER BY date
            """, (cutoff,))
            return cursor.fetchall()
        except sqlite3.Error:
            return []

    # ── 엔진 상태 공유 (웹 모니터용) ──────────────────────────
    def export_to_excel(self, output_path: str, days: int = 90) -> bool:
        """매매 기록을 엑셀 파일로 내보내기."""
        if not EXCEL_AVAILABLE:
            logger.error("❌ [DB] openpyxl 미설치 — pip install openpyxl")
            return False
        try:
            cutoff = (datetime.datetime.now() - datetime.timedelta(days=days)).strftime("%Y-%m-%d")
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT date, time, trade_type, condition_name, stock_name, stock_code,
                       exec_price, exec_qty, realized_profit, commission, tax, order_type,
                       CASE WHEN is_mock=1 THEN '모의' ELSE '실계좌' END, sell_reason
                FROM trade_history WHERE date >= ? ORDER BY date, time
            """, (cutoff,))
            rows = cursor.fetchall()

            wb = openpyxl.Workbook()

            # ── 시트 1: 전체 매매 기록 ──
            ws = wb.active
            ws.title = "매매기록"
            headers = ["날짜", "시간", "매매구분", "조건식", "종목명", "종목코드",
                       "체결가", "수량", "실현손익", "수수료", "세금", "주문유형", "계좌구분", "매도사유"]
            header_fill = PatternFill("solid", fgColor="1E3A5F")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            profit_font_pos = Font(color="FF4444", bold=True)
            profit_font_neg = Font(color="4488FF", bold=True)
            buy_fill = PatternFill("solid", fgColor="1A2A1A")
            sell_fill = PatternFill("solid", fgColor="1A1A2A")

            for r_idx, row in enumerate(rows, 2):
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                    # 숫자 컬럼 포맷
                    if c_idx in (7, 9, 10, 11):  # 체결가, 실현손익, 수수료, 세금
                        cell.number_format = "#,##0"
                    # 실현손익 색상
                    if c_idx == 9 and val:
                        cell.font = profit_font_pos if val > 0 else (profit_font_neg if val < 0 else Font())
                    # 행 배경
                    if row[2] == "매수":
                        cell.fill = buy_fill
                    else:
                        cell.fill = sell_fill

            # 컬럼 너비 자동 조정
            col_widths = [12, 10, 10, 18, 14, 10, 12, 8, 14, 12, 12, 12, 10, 16]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # ── 시트 2: 일별 손익 ──
            ws2 = wb.create_sheet("일별손익")
            ws2.append(["날짜", "실현손익", "매도건수", "누적손익"])
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            cumulative = 0
            for date, pnl, cnt in self.get_daily_pnl(days=days):
                cumulative += pnl
                row_data = [date, pnl, cnt, cumulative]
                ws2.append(row_data)
                r = ws2.max_row
                ws2.cell(r, 2).number_format = "#,##0"
                ws2.cell(r, 4).number_format = "#,##0"
                ws2.cell(r, 2).font = profit_font_pos if pnl > 0 else profit_font_neg
            for col in ["A", "B", "C", "D"]:
                ws2.column_dimensions[col].width = 14

            # ── 시트 3: 조건식별 성과 ──
            ws3 = wb.create_sheet("조건식성과")
            ws3.append(["조건식", "매도건수", "승수", "총손익", "승률(%)"])
            for cell in ws3[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for cond_name, cnt, wins, total_pnl in self.get_condition_performance(days=days):
                win_rate = round(wins / cnt * 100, 1) if cnt else 0
                ws3.append([cond_name, cnt, wins, total_pnl, win_rate])
                r = ws3.max_row
                ws3.cell(r, 4).number_format = "#,##0"
            for col, w in zip(["A", "B", "C", "D", "E"], [20, 12, 10, 14, 12]):
                ws3.column_dimensions[col].width = w

            wb.save(output_path)
            logger.info(f"📊 [DB] 엑셀 내보내기 완료: {output_path} ({len(rows)}건)")
            return True
        except Exception as e:
            logger.error(f"❌ [DB] 엑셀 내보내기 실패: {e}")
            return False

    def write_engine_state(self, state: dict):
        """[개선 #4] 엔진 상태를 JSON 파일로 저장 (웹 모니터가 읽음). 원자적 쓰기."""
        tmp_path = self._state_file + ".tmp"
        try:
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(state, f, ensure_ascii=False)
            os.replace(tmp_path, self._state_file)
        except Exception as e:
            logger.debug(f"[DB] 상태 파일 쓰기 실패: {e}")

    def read_engine_state(self) -> dict:
        """[개선 #4] 엔진 상태 JSON 파일 읽기 (웹 모니터용)."""
        try:
            if os.path.exists(self._state_file):
                with open(self._state_file, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception as e:
            logger.debug(f"[DB] 상태 파일 읽기 실패: {e}")
        return {}

    def close(self):
        if self.conn:
            self.conn.close()
